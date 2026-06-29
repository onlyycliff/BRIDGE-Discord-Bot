from flask import Flask, render_template, request, jsonify
from bridge_bot.bot import send_poll, start_bot, bot
from bridge_bot.excel_manager import excel_manager
import threading
import asyncio
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
import csv
from io import StringIO
from pathlib import Path

dashboard = Flask(__name__)

# Configuration
BASE_DIR = Path(__file__).parent
EXCEL_FILE = BASE_DIR / 'responses.xlsx'

# Routes
@dashboard.route('/')
def home():
    """Serve the main dashboard"""
    return render_template('dashboard.html')

@dashboard.route('/submit', methods=["POST"])
def submit():
    """Handle poll submission from frontend"""
    try:
        data = request.get_json()
        if not data or 'question' not in data:
            return jsonify({"error": "Invalid data"}), 400
        
        question = data.get("question")
        option1 = data.get("option1")
        option2 = data.get("option2")
        
        if not all([question, option1, option2]):
            return jsonify({"error": "Missing required fields"}), 400
        
        bot.loop.create_task(send_poll(question, option1, option2))
        return jsonify({"message": "Poll created successfully", "status": "ok"})
    except Exception as e:
        print(f"Error submitting poll: {e}")
        return jsonify({"error": str(e)}), 500

# API Endpoints

@dashboard.route('/api/polls', methods=['GET'])
def get_polls():
    """Return all polls with their vote counts"""
    try:
        if not EXCEL_FILE.exists():
            return jsonify([])
        
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        polls = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            
            timestamp, username, question, choice = row[:4]
            
            if question not in polls:
                polls[question] = {
                    'question': question,
                    'options': {},
                    'timestamp': timestamp
                }
            
            if choice not in polls[question]['options']:
                polls[question]['options'][choice] = {'name': choice, 'votes': 0}
            
            polls[question]['options'][choice]['votes'] += 1
        
        result = []
        for poll in polls.values():
            poll['options'] = list(poll['options'].values())
            result.append(poll)
        
        return jsonify(result)
    except Exception as e:
        print(f"Error loading polls: {e}")
        return jsonify({"error": str(e)}), 500

@dashboard.route('/api/votes', methods=['GET'])
def get_votes():
    """Return all votes from Excel"""
    try:
        if not EXCEL_FILE.exists():
            return jsonify([])
        
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        votes = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            
            timestamp, username, question, choice = row[:4]
            votes.append({
                'timestamp': timestamp,
                'username': username,
                'question': question,
                'choice': choice
            })
        
        return jsonify(votes)
    except Exception as e:
        print(f"Error loading votes: {e}")
        return jsonify({"error": str(e)}), 500

@dashboard.route('/api/votes/export', methods=['GET'])
def export_votes():
    """Export votes as CSV"""
    try:
        if not EXCEL_FILE.exists():
            return '', 204
        
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['Timestamp', 'Username', 'Question', 'Choice'])
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            writer.writerow(row[:4])
        
        response_text = output.getvalue()
        return response_text, 200, {
            'Content-Disposition': 'attachment; filename=vote-log.csv',
            'Content-Type': 'text/csv'
        }
    except Exception as e:
        print(f"Error exporting votes: {e}")
        return jsonify({"error": str(e)}), 500

@dashboard.route('/api/bot-status', methods=['GET'])
def get_bot_status():
    """Return bot status information"""
    try:
        # Calculate bot stats from Excel
        votes_today = 0
        votes_total = 0
        
        if EXCEL_FILE.exists():
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            today = datetime.now().date()
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    votes_total += 1
                    if isinstance(row[0], datetime):
                        if row[0].date() == today:
                            votes_today += 1
        
        return jsonify({
            'online': True,
            'uptime': '2d 14h 32m',
            'last_command': '/poll',
            'votes_today': votes_today,
            'votes_total': votes_total
        })
    except Exception as e:
        print(f"Error getting bot status: {e}")
        return jsonify({"error": str(e)}), 500

@dashboard.route('/api/schedule', methods=['GET'])
def get_schedule():
    """Return workshop schedule"""
    try:
        schedule = [
            {
                'id': 1,
                'name': 'Opening Keynote',
                'pillar': 'Professional Development',
                'speaker': 'Dr. Jane Smith',
                'time': '9:00 AM - 10:00 AM',
                'location': 'Main Hall',
                'description': 'Welcome to Bridge 2026! Learn about our mission and goals for the year.'
            },
            {
                'id': 2,
                'name': 'Holistic Engineering Workshop',
                'pillar': 'Holistic STEMinist',
                'speaker': 'Prof. Ahmed Hassan',
                'time': '10:30 AM - 12:00 PM',
                'location': 'Room 201',
                'description': 'Explore the intersection of engineering, ethics, and social impact.'
            },
            {
                'id': 3,
                'name': 'Career Development Panel',
                'pillar': 'Professional Development',
                'speaker': 'Industry Leaders',
                'time': '1:00 PM - 2:30 PM',
                'location': 'Main Hall',
                'description': 'Hear from successful professionals about career paths and opportunities.'
            },
            {
                'id': 4,
                'name': 'STEMinist Roundtable',
                'pillar': 'Holistic STEMinist',
                'speaker': 'Student Leaders',
                'time': '3:00 PM - 4:30 PM',
                'location': 'Room 105',
                'description': 'Interactive discussion on diversity and inclusion in STEM.'
            }
        ]
        return jsonify(schedule)
    except Exception as e:
        print(f"Error getting schedule: {e}")
        return jsonify({"error": str(e)}), 500

# New Excel Manager Endpoints

@dashboard.route('/api/excel/stats', methods=['GET'])
def get_excel_stats():
    """Get statistics from Excel data"""
    try:
        all_votes = excel_manager.get_all_votes()
        
        if not all_votes:
            return jsonify({
                'total_votes': 0,
                'unique_voters': 0,
                'polls_count': 0,
                'data_status': 'No data yet'
            })
        
        unique_voters = len(set(v['User_ID'] for v in all_votes if 'User_ID' in v))
        polls_count = len(set(v.get('Poll_ID') for v in all_votes if 'Poll_ID' in v))
        
        return jsonify({
            'total_votes': len(all_votes),
            'unique_voters': unique_voters,
            'polls_count': polls_count,
            'data_status': 'Synced',
            'last_updated': datetime.now().isoformat()
        })
    except Exception as e:
        print(f"Error getting Excel stats: {e}")
        return jsonify({"error": str(e)}), 500

@dashboard.route('/api/excel/summary', methods=['GET'])
def get_excel_summary():
    """Get summary statistics by question"""
    try:
        summary = excel_manager.get_summary_by_question()
        
        result = []
        for question, data in summary.items():
            result.append({
                'question': question,
                'total_votes': int(data.get('Total_Votes', 0)),
                'choices': data.get('Choice', {})
            })
        
        return jsonify(result)
    except Exception as e:
        print(f"Error getting Excel summary: {e}")
        return jsonify({"error": str(e)}), 500

@dashboard.route('/api/excel/export-csv', methods=['GET'])
def export_excel_csv():
    """Export all Excel data to CSV"""
    try:
        csv_path = excel_manager.export_to_csv()
        
        if not csv_path or not Path(csv_path).exists():
            return jsonify({"error": "Export failed"}), 500
        
        with open(csv_path, 'r') as f:
            csv_content = f.read()
        
        return csv_content, 200, {
            'Content-Disposition': 'attachment; filename=bridge-2026-feedback.csv',
            'Content-Type': 'text/csv'
        }
    except Exception as e:
        print(f"Error exporting CSV: {e}")
        return jsonify({"error": str(e)}), 500

@dashboard.route('/api/excel/poll/<poll_id>', methods=['GET'])
def get_poll_details(poll_id):
    """Get details for a specific poll"""
    try:
        stats = excel_manager.get_poll_stats(poll_id)
        
        if not stats:
            return jsonify({"error": "Poll not found"}), 404
        
        return jsonify(stats)
    except Exception as e:
        print(f"Error getting poll details: {e}")
        return jsonify({"error": str(e)}), 500

def run_bot():
    """Start the Discord bot in a separate thread"""
    start_bot()













if __name__ == '__main__':
    bot_thread = threading.Thread(target=run_bot, daemon=True)
    bot_thread.start()
    dashboard.run(debug=False, port=5000)